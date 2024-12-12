import requests

#無料で〒を取得するAPI
url3 = "https://api.excelapi.org/post/zipcode"

#３：新規でエクセルを開き、取得したデータを書き込んで名前を付けて保存する
from spire.xls import Workbook
from spire.xls import Worksheet

# セル幅調整し、データ保存し、データ完成
import openpyxl

#検索したい店舗情報をデータに格納していく
wb1 = openpyxl.load_workbook("shoplist.xlsx")
ws1 = wb1.worksheets[0]

# 最大行番号を取得
max = ws1.max_row

# 最終行に値があるか確認する
while True:
    if ws1.cell(row=max, column=1).value == None:
        max -= 1
    else:
        break

kw = []
    
for row in ws1["A1:A"+str(max)]:
    for col in row:
        kw.append(col.value)
print(kw)    

wb1.close()

#１：無料のホットペッパーAPIで店舗名からデータを取得
URL = 'http://webservice.recruit.co.jp/hotpepper/gourmet/v1/'
API_KEY = '' #　←　ここの''の中にホットペッパーAPIのキーを入力が必要

index = 0

# Excelブックを新規で開く
workbook = Workbook()

# ワークシートを削除して新しいワークシートを作成
worksheet = workbook.Worksheets.Add("店舗情報")

# エクセルシートにタイトルを付ける
worksheet.Range["A1"].Text = "店舗名"
worksheet.Range["B1"].Text = "住所"
worksheet.Range["C1"].Text = "郵便番号"

for KEYWORD in kw:
    params = {
        'key':API_KEY,
        'keyword':KEYWORD,
        'format':'json',
        'count':1
    }

    response = requests.get(URL,params)
    data = response.json()

    #（正式な）店名と住所を取得
    for shop in data['results']['shop']:
        name = shop['name']
        address = shop['address']
        
        print(f"店名: {name}, 住所: {address}")
        
        A = (2 + index)
        B = (2 + index)
        C = (2 + index)

        # pythonからエクセルへデータを書く
        worksheet.Range["A" + str(A)].Text = name
        worksheet.Range["B" + str(B)].Text = address
        text = "\"https://api.excelapi.org/post/zipcode?address=\""
        text2 = "=WEBSERVICE(" + text +"&ENCODEURL("+"B" + str(B)+"))"
        #print(text2)
        worksheet.Range["C" + str(C)].Text = text2
        index = index + 1

# ワークブックをいったん保存
workbook.SaveToFile("output/店舗データ.xlsx")
workbook.Dispose()
        
wb = openpyxl.load_workbook('output/店舗データ.xlsx')
ws = wb['店舗情報']

for col in ws.columns:
    max_length = 0
    column = col[0].column_letter
    for cell in col:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(cell.value)
        except:
            pass
    adjusted_width = (max_length + 1) * 2
    ws.column_dimensions[column].width = adjusted_width

wb.save('output/店舗データ.xlsx')

# 終了時にWindowsがメッセージを表示    
import tkinter as tk
import tkinter.messagebox as messagebox

tk.Tk().withdraw()
messagebox.showinfo('メッセージ', '処理が完了しました。')
